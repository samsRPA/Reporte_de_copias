import os
import re
import pandas as pd 
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def dejar_archivo():
    # Ruta donde están los archivos
    folder_path = r"C:\\Users\\User\\Downloads"

    # Expresión regular para extraer la fecha y hora del nombre del archivo
    pattern = r"NUEVA CONSULTA INFORME COPIAS CON GESTION NOTIFICADORES_(\w{3})([A-Za-z]{3})(\d{2})(\d{6})COT(\d{4})\.csv"

    # Obtener los archivos de la carpeta
    files = os.listdir(folder_path)

    # Filtrar solo los archivos que coinciden con el patrón
    csv_files = [f for f in files if re.match(pattern, f)]

    print("Archivos encontrados:", csv_files)  # Depuración

    # Lista para almacenar los archivos con su fecha y hora extraída
    file_info = []

    for file in csv_files:
        match = re.search(pattern, file)
        if match:
            day_of_week = match.group(1)  # Día de la semana (Ej: Mon, Tue)
            month_str = match.group(2)  # Mes (Ej: Feb)
            day = match.group(3)  # Día del mes (Ej: 04)
            time_part = match.group(4)  # Hora (Ej: 171741)
            year = match.group(5)  # Año (Ej: 2025)

            # Crear una fecha en formato adecuado para ordenación
            full_date_str = f"{day} {month_str} {year} {time_part}"  # Ej: "04 Feb 2025 171741"
            try:
                file_datetime = datetime.strptime(full_date_str, '%d %b %Y %H%M%S')
                file_info.append((file, file_datetime))
            except ValueError as e:
                print(f"Error en conversión de fecha para {file}: {e}")

    # Ordenar los archivos por fecha y hora (del más reciente al más antiguo)
    file_info.sort(key=lambda x: x[1], reverse=True)

    if not file_info:
        print("No se encontraron archivos válidos.")
        return None

    # El archivo más reciente
    most_recent_file = file_info[0][0]

    # Eliminar archivos más antiguos
    for file, _ in file_info[1:]:
        file_path = os.path.join(folder_path, file)
        os.remove(file_path)
        print(f"Archivo eliminado: {file}")

    print(f"El archivo más reciente es: {most_recent_file}")

    return os.path.join(folder_path, most_recent_file)


def mover_columnas(df):
    """ MOVER COLUMNAS AL LADO IZ DE COPIAS """
    # Existencia columnas
    if 'COPIAS' in df.columns and 'ACTUACION_NOMBRE' in df.columns and 'NOTIFICACION_NOMBRE' in df.columns:
        # Crear una lista del orden deseado de columnas
        columnas = df.columns.tolist()
        
        # Identificar índices relevantes
        indice_copias = columnas.index('COPIAS')
        
        # Quitar las columnas que queremos mover
        columnas.remove('ACTUACION_NOMBRE')
        columnas.remove('NOTIFICACION_NOMBRE')
        
        # Insertar las columnas movidas después de 'COPIAS'
        columnas[indice_copias + 1:indice_copias + 1] = ['ACTUACION_NOMBRE', 'NOTIFICACION_NOMBRE']
        
        # Reordenar el DataFrame
        df = df[columnas]
    else:
        print("Alguna de las columnas no existe en el DataFrame.")
    
    return df

def mover_columna_rad(df):
    """Mover la columna 'RADICACION' a la primera posición."""
    if 'RADICACION' in df.columns:
        # Crear una lista de columnas
        columnas = df.columns.tolist()
        
        # Mover 'RADICACION' al inicio de la lista
        columnas.remove('RADICACION')  # Eliminar 'RADICACION' de su posición actual
        columnas.insert(0, 'RADICACION')  # Insertarla al inicio
        
        # Reordenar el DataFrame
        df = df[columnas]
    else:
        print("La columna 'RADICACION' no existe en el DataFrame.")
    
    return df

def mover_columna_despacho(df):
    """ MOVER COLUMNA AL LADO DR COLUMNA GESTION_COPIA """
    # Existencia columnas
    if 'DESPACHO_ID' in df.columns and 'GESTION_COPIA' in df.columns:
        # Crear una lista del orden actual de columnas
        columnas = df.columns.tolist()
        
        # Identificar el índice de 'GESTION_COPIA'
        indice_gestion_copia = columnas.index('GESTION_COPIA')
        
        # Quitar la columna 'DESPACHO_ID' de su posición actual
        columnas.remove('DESPACHO_ID')
        
        # Insertar 'DESPACHO_ID' justo después de 'GESTION_COPIA'
        columnas.insert(indice_gestion_copia + 1, 'DESPACHO_ID')
        
        # Reordenar el DataFrame con las columnas modificadas
        df = df[columnas]
        print("\nColumna 'DESPACHO_ID' movida al lado derecho de 'GESTION_COPIA'.")
    else:
        print("Alguna de las columnas no existe en el DataFrame.")
    
    return df

def leer_archivo():
    archivo_reciente = dejar_archivo()

    textos_a_eliminar = [
        "NO LISTADO",
        "RADICACION",
        "NOTIFICACION POR ESTRADO",
        "NOTIFICACION CLIENTE",
        "ACT SECRETARIAL",
        "AUTOS",
        "RAMA JUDICIAL",
        "AVISO DE APERTURA",
        "REQUERIR",
        "GESTION CLIENTE",
        "AUTOS DE CUMPLASE",
        "RAMA JUDICIAL - NO GENERA COPIA"
        "DESCARGA DE DOCUMENTOS SIUGJ",
        "ENTRADAS AL DESPACHO",
        "VISITA A JUZGADOS",
        "ACT SECRETARIAL",
        "AVISO DE SALA",
        "AVISOS DE REMATE",
    ]
    
    try:
        fecha_actual = datetime.now().strftime('%d-%m-%Y')

        df = pd.read_csv(archivo_reciente, encoding='latin1', sep=';', low_memory=False)

        # Cambiar el nombre del archivo para guardarlo en la carpeta especificada
        archivo_xlsx = rf"C:\Users\User\OneDrive - LITIGAR PUNTO COM S.A\INFORME COPIAS\informe_copias_{fecha_actual}.xlsx"

        # Eliminar la columna 'GESTION_COPIA' si existe
        if 'VALOR_MANDAMIENTO' in df.columns:
            df = df.drop(columns=['VALOR_MANDAMIENTO'])
            
        # Eliminar la columna 'NOMBRE_DEPENDIENTE' si existe
        if 'NOMBRE_DEPENDIENTE' in df.columns:
            df = df.drop(columns=['NOMBRE_DEPENDIENTE'])
        # Eliminar las filas de la columna 'COPIAS' que contienen el texto 'CONCOPIA'
        if 'COPIAS' in df.columns:
            df = df[df['COPIAS'] != 'CONCOPIA']
        # Eliminar las filas donde 'GESTION_COPIA' no es NaN
        df = df[df['GESTION_COPIA'].isna()]

        # Eliminar las filas donde 'ORIGEN' sea igual a 'OPERACION'
        df = df[df['ORIGEN'] != 'OPERAC']
        
        # Eliminar los duplicados de la columna 'ACTUACION_PROCESAL_ID'
        df = df.drop_duplicates(subset=['ACTUACION_PROCESAL_ID'])
        
        # Eliminar las filas donde 'NOTIFICACION_NOMBRE' contenga alguno de los textos especificados
        df = df[~df['NOTIFICACION_NOMBRE'].isin(textos_a_eliminar)]
        
        # Mover las columnas 
        df = mover_columnas(df)
        
        # Mover columna radicadcion 
        df = mover_columna_rad(df)

        # Mover columna despacho
        df = mover_columna_despacho(df)
        print(df)
        # Ordenar por 'NOTIFICADOR_NOMBRE', luego por 'FECHA_NOTIFICACION', luego por 'DESPACHO_ID'
        if all(col in df.columns for col in ['NOMBRE_NOTIFICADOR', 'FECHA_NOTIFICACION', 'DESPACHO_ID']):

            def parse_fecha(fecha):
                from pandas.errors import OutOfBoundsDatetime
                import pandas as pd

                # Intenta formatos conocidos
                formatos = ['%d/%m/%Y', '%m/%d/%Y %I:%M:%S %p']

                for fmt in formatos:
                    try:
                        return pd.to_datetime(fecha, format=fmt, errors='raise')
                    except (ValueError, TypeError, OutOfBoundsDatetime):
                        continue
                # Si no pudo con formatos exactos, intenta parseo flexible con dayfirst=True
                try:
                    return pd.to_datetime(fecha, errors='coerce', dayfirst=True)
                except:
                    return pd.NaT

            df['FECHA_NOTIFICACION'] = df['FECHA_NOTIFICACION'].apply(parse_fecha)

            df = df.sort_values(by=['NOMBRE_NOTIFICADOR', 'FECHA_NOTIFICACION', 'DESPACHO_ID'])
            print("\nDatos ordenados por 'NOMBRE_NOTIFICADOR', 'FECHA_NOTIFICACION' y 'DESPACHO_ID'.")
        
        # Guardar el archivo como Excel
        df.to_excel(archivo_xlsx, index=False, engine='openpyxl')
        print(f"\nArchivo guardado como Excel: {archivo_xlsx}")

        # Aplicar formato al archivo Excel
        wb = load_workbook(archivo_xlsx)
        ws = wb.active

        # Aplicar color amarillo a la columna 'GESTION_COPIA'
        if 'GESTION_COPIA' in df.columns:
            col_index = df.columns.get_loc('GESTION_COPIA') + 1  # Obtener índice (1-based)
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            for row in range(2, ws.max_row + 1):  # Desde la fila 2 (omitir encabezado)
                ws.cell(row=row, column=col_index).fill = yellow_fill

        # Ajustar el alto de todas las filas a 15
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 15

        # Guardar los cambios en el archivo Excel
        wb.save(archivo_xlsx)
        print("\nFormato aplicado al archivo Excel.")

        return df
    
    except Exception as e:
        print(f"Error al leer el archivo CSV: {e}")
    
# Llamada a la función para leer el archivo más reciente
leer_archivo()