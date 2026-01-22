from openpyxl import load_workbook
from openpyxl.styles import Font
from pathlib import Path
import logging
from datetime import datetime
import subprocess
from app.domain.interfaces.IProcessDataService import IProcessDataService
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.page import PageMargins
from PIL import Image
import base64
from io import BytesIO

from openpyxl.worksheet.worksheet import Worksheet
class ProcessDataService(IProcessDataService):

    def __init__(self):
        self.logger = logging.getLogger(__name__)

    def generate_monitoring_report_xlsx( self,monitoring_report: list,column_names: list):
        try: 
            # ==========================
            # RUTAS
            # ==========================
            base_template = Path("/app/output/base/INFORME_DRA_ORIGEN.xlsx")

            fecha = datetime.now().strftime("%d-%m-%Y")
            excel_path = Path(f"/app/output/reports/INFORME_DRA_{fecha}.xlsx")
            excel_path.parent.mkdir(parents=True, exist_ok=True)

            # ==========================
            # CARGAR EXCEL BASE
            # ==========================
            wb = load_workbook(base_template)
            sheet_name = "BASE DE DATOS DESCARGAS"

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"La hoja '{sheet_name}' no existe en el Excel base")

            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row)

            # ==========================
            # ENCABEZADOS (BD)
            # ==========================
            headers = [col[0] for col in column_names]

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)

            # ==========================
            # DATA
            # ==========================
            for row in monitoring_report:
                ws.append(list(row))

            wb.save(excel_path)
            self.logger.info(f"📊 Excel generado → {excel_path}")
            
            return excel_path
        
        except Exception as e:
            self.logger.error(f"❌ Error : {e}")
            return None

    # ==========================================================
    # AJUSTAR HOJA 1 A HORIZONTAL (SIN MACROS)
    # ==========================================================


    def _prepare_sheet(self, excel_path: Path):
        """
        Fuerza área A1:P31 y agranda el TAMAÑO DE PÁGINA
        para que LibreOffice NO recorte contenido.
        """

        wb = load_workbook(excel_path)
        ws: Worksheet = wb.worksheets[0]

        # ==========================
        # ORIENTACIÓN HORIZONTAL
        # ==========================
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

        # ==========================
        # 🔥 TAMAÑO DE PAPEL GRANDE
        # ==========================
        # 8 = A3, 9 = A4, 11 = A2
        ws.page_setup.paperSize = ws.PAPERSIZE_A3  # ← CLAVE
        # Si A3 aún queda justo, usa:
        # ws.page_setup.paperSize = ws.PAPERSIZE_A2

        # ==========================
        # ÁREA EXACTA
        # ==========================
        ws.print_area = "A1:P31"

        # ==========================
        # AJUSTE A UNA PÁGINA
        # ==========================
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.scale = 91

        # ==========================
        # MÁRGENES 
        # ==========================
        ws.page_margins = PageMargins(
            left=0.3,
            right=0,
            top=0.5,
            bottom=0
        )

        wb.save(excel_path)

      

    # ==========================================================
    # EXCEL → IMAGEN
    # ==========================================================

    def capture_img(self, excel_path: Path):
        """
        Convierte la PRIMERA hoja del Excel directamente a imagen (PNG),
        la recorta y retorna:
        - image_path (Path)
        - image_base64 (str)
        """
        # ==========================
        # AJUSTAR HOJA 1
        # ==========================
        self._prepare_sheet(excel_path)

        fecha = datetime.now().strftime("%d-%m-%Y")
        output_dir = excel_path.parent
        image_path = output_dir / f"captura_informe_{fecha}.png"

        try:
            # ==========================
            # CONVERTIR EXCEL → PNG
            # ==========================
            subprocess.run(
                [
                    "libreoffice",
                    "--headless",
                    "--calc",
                    "--convert-to",
                    "png:calc_png_Export:Resolution=300",
                    "--outdir",
                    str(output_dir),
                    str(excel_path)
                ],
                check=True
            )

            generated_image = output_dir / f"{excel_path.stem}.png"
            generated_image.rename(image_path)

            self.logger.info(f"🖼️ Imagen generada → {image_path}")

            # ==========================
            # RECORTAR + BASE64
            # ==========================
            with Image.open(image_path) as img:
                width, height = img.size

                cropped_img = img.crop(
                    (0, 0, width, int(height / 1.9))
                )

                # Guardar imagen final
                cropped_img.save(image_path, format="PNG")

                # Convertir a Base64
                buffer = BytesIO()
                cropped_img.save(buffer, format="PNG")
                image_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")

            self.logger.info("✅ Imagen procesada y convertida a Base64")

            return image_path, image_base64

        except subprocess.CalledProcessError as e:
            self.logger.error("❌ Error convirtiendo Excel a imagen", exc_info=e)
            raise
