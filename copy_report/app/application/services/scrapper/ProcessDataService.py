from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pathlib import Path
import logging
from datetime import datetime

class ProcessDataService:

    logger = logging.getLogger(__name__)

    TEXTOS_A_ELIMINAR = {
        "NO LISTADO",
        "RADICACION",
        "NOTIFICACION POR ESTRADO",
        "NOTIFICACION CLIENTE",
        "ACT SECRETARIAL",
        "AUTOS",
        "RAMA JUDICIAL",
        "AVISO DE APERTURA",
        "REQUERIR",
        "GESTION CLIENTE",
        "AUTOS DE CUMPLASE",
        "RAMA JUDICIAL - NO GENERA COPIA",
        "DESCARGA DE DOCUMENTOS SIUGJ",
        "ENTRADAS AL DESPACHO",
        "VISITA A JUZGADOS",
        "AVISO DE SALA",
        "AVISOS DE REMATE",
    }

    def generate_copy_report_xlsx(self, copy_report: list, output_path: str):

        # ==========================
        # 1. FILTRADO (equivalente a pandas)
        # ==========================
        filtrado = []
        vistos = set()  # para duplicados por ACTUACION_PROCESAL_ID

        for item in copy_report:

            if item.copias == "CONCOPIA":
                continue

            if item.gestion_copia is not None:
                continue

            if item.origen == "OPERAC":
                continue

            if item.notifiacion_nombre in self.TEXTOS_A_ELIMINAR:
                continue

            if item.actuacion_procesal_id in vistos:
                continue

            vistos.add(item.actuacion_procesal_id)
            filtrado.append(item)

        # ==========================
        # 2. ORDENAMIENTO
        # ==========================
        def parse_fecha(fecha):
            if not fecha:
                return datetime.min
            if isinstance(fecha, datetime):
                return fecha
            try:
                return datetime.strptime(fecha, "%d/%m/%Y")
            except:
                try:
                    return datetime.strptime(fecha, "%m/%d/%Y %I:%M:%S %p")
                except:
                    return datetime.min

        filtrado.sort(
            key=lambda x: (
                x.nombre_notificador or "",
                parse_fecha(x.fecha_notificacion),
                x.despacho_id or 0
            )
        )

        # ==========================
        # 3. CREACIÓN EXCEL
        # ==========================
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "CopyReport"

        headers = [
            "RADICACION",
            "PROCESO_ID",
            "ACTUACION_PROCESAL_ID",
            "NOTIFICACION_NOMBRE",
            "RESUELVE",
            "COPIAS",
            "GESTION_COPIA",
            "FECHA_NOTIFICACION",
            "DESPACHO_NOMBRE",
            "CIUDAD",
            "COORDINADOR",
            "ORIGEN",
            "REGIONAL",
            "NOMBRE_NOTIFICADOR",
            "TIPO_PROCESO",
            "AÑO",
            "DEMANDANTE",
            "DEMANDADO",
            "LOCALIDAD_ID",
            "DESPACHO_ID",
            "INSTANCIA_RADICACION",
            "ACTUACION_FECHA",
            "ACTUACION_FECHA_AUTO",
            "ACTUACION_FECHA_CARGA",
            "ETAPA_NOMBRE",
            "ACTUACION_NOMBRE",
            "CLIENTE_ID",
            "TIPO_CLIENTE",
            "CLIENTE_NOMBRE",
            "GESTOR",
            "AUTORIZACION",
        ]

        # Encabezados
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # Datos
        for item in filtrado:
            ws.append([
                item.radicacion,
                item.proceso_id,
                item.actuacion_procesal_id,
                item.notifiacion_nombre,
                item.resuelve,
                item.copias,
                item.gestion_copia,
                item.fecha_notificacion,
                item.despacho_nombre,
                item.ciudad,
                item.coordinador,
                item.origen,
                item.regional,
                item.nombre_notificador,
                item.tipo_proceso,
                item.annio,
                item.demandante,
                item.demandado,
                item.localidad_id,
                item.despacho_id,
                item.instancia_radicacion,
                item.actuacion_fecha,
                item.actuacion_fecha_auto,
                item.actuacion_fecha_carga,
                item.etapa_nombre,
                item.actuacion_nombre,
                item.cliente_id,
                item.tipo_cliente,
                item.cliente_nombre,
                item.gestor,
                item.autorizacion,
            ])

        # ==========================
        # 4. FORMATO EXCEL
        # ==========================
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        gestion_col = headers.index("GESTION_COPIA") + 1

        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=gestion_col).fill = yellow_fill
            ws.row_dimensions[row].height = 15

        wb.save(output_path)

        self.logger.info(f"📊 XLSX generado correctamente → {output_path}")
